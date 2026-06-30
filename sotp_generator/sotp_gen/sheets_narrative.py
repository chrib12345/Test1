"""Narrative / documentation sheets: Cover, Build Spec, Spin Mechanics,
Recent Trends, Street Coverage. Driven by config text with templated defaults;
the Cover and Street/Summary cross-links stay wired to the engine.
"""
from openpyxl.utils import get_column_letter
from . import styles as ST
from .styles import cell, section, input_cell, note


def build_cover(ws, c):
    A, S = c["A"], c["S"]
    meta = c["meta"]
    ST.setup_sheet(ws, ST.NAVY, widths={"A": 2, "B": 40, "C": 16, "D": 58, "E": 12, "H": 24})
    cell(ws, "B3", f"{meta.get('company','')}  ({meta.get('exchange','')}: {meta.get('ticker','')})", bold=True)
    cell(ws, "B4", meta.get("subtitle", "SUM-OF-THE-PARTS VALUATION"), bold=True)
    cell(ws, "B5", meta.get("built_on", "Built on master SOTP/Spin template  |  "
                            "Refresh blue inputs before use."), font=ST.FONT_NOTE)

    section(ws, 8, "OUTPUT SUMMARY", last_col="D")
    out = [
        ("Current share price", f"=Assumptions!B{A['price']}", ST.FMT_USD_2),
        ("SOTP — Bear", f"=SOTP!B{S['per_share']}", ST.FMT_USD_2),
        ("SOTP — Base (peer-comp multiples)", f"=SOTP!C{S['per_share']}", ST.FMT_USD_2),
        ("SOTP — Re-Rate", f"=SOTP!D{S['per_share']}", ST.FMT_USD_2),
        ("Base-case upside / (downside)", f"=SOTP!C{S['upside']}", ST.FMT_PCT),
    ]
    if c["spin"].get("enabled"):
        out += [("Base case — RemainCo per share", f"=SOTP!B{S['remain_ps']}", ST.FMT_USD_2),
                (f"Base case — {c['spin'].get('entity_ticker','spin')} per share", f"=SOTP!B{S['spin_ps']}", ST.FMT_USD_2)]
    for j, (lab, f, fmt) in enumerate(out):
        rr = 9 + j
        cell(ws, f"B{rr}", lab)
        cell(ws, f"C{rr}", f, fmt=fmt, font=ST.FONT_LINK)

    base_guide = 9 + len(out) + 2
    section(ws, base_guide, "TAB GUIDE", last_col="D")
    guide = [
        ("Build Spec", "READ FIRST — research sequence, data rules, net-debt protocol, verification"),
        ("Assumptions", "Segment financials, corporate bridge, multiples, net-debt build & NCI, sources"),
        ("SOTP", "EBITDA build, Bear/Base/Re-Rate bridge, entity split, reverse SOTP"),
        ("Valuation Summary", "Scenario table and football-field ranges"),
        ("FCF Build", "EBITDA-to-FCF bridge, leverage & capital-return capacity"),
        ("Sensitivity", "Heat-mapped multiple grids — price and upside"),
        ("Comps", "Graded peers per segment, live EV build, regression"),
        ("Recent Trends", "Quarterly trend and peer trends"),
        ("Street Coverage", "Targets, consensus, model-vs-Street variance, catalysts"),
        ("Spin Mechanics", "Deal facts, timeline, key judgment calls"),
    ]
    for j, (tab, desc) in enumerate(guide):
        rr = base_guide + 1 + j
        cell(ws, f"B{rr}", tab, bold=True)
        cell(ws, f"D{rr}", desc, font=ST.FONT_NOTE, wrap=True)

    base_meth = base_guide + len(guide) + 2
    section(ws, base_meth, "METHODOLOGY", last_col="D")
    cell(ws, f"B{base_meth+1}", meta.get("methodology",
         "Each segment valued on EV / NTM standalone Adj. EBITDA via peer-comp multiples. "
         "Segment Adj. EBITDA = adjusted operating profit + allocated depreciation. Corporate cost "
         "is capitalized at the blended multiple and deducted. Any segment minority is deducted at "
         "fair value. Net debt is built from the latest balance sheet and marked to the spin date."),
         font=ST.FONT_BASE, wrap=True)
    ws.merge_cells(f"B{base_meth+1}:H{base_meth+1}")
    cell(ws, f"B{base_meth+3}", "For analytical purposes only — not investment advice. "
         "Refresh blue inputs (prices, peer fundamentals, estimates) before use.", font=ST.FONT_WARN)

    calls = meta.get("judgment_calls", [])
    if calls:
        base_jc = base_meth + 5
        section(ws, base_jc, "KEY JUDGMENT CALLS (see Spin Mechanics)", last_col="D")
        for j, txt in enumerate(calls):
            rr = base_jc + 1 + j
            cell(ws, f"B{rr}", f"{j+1}. {txt}", font=ST.FONT_NOTE, wrap=True)
            ws.merge_cells(f"B{rr}:H{rr}")


def build_buildspec(ws, c):
    ST.setup_sheet(ws, ST.WARN_RED if False else "C00000", widths={"A": 4, "B": 120})
    cell(ws, "A1", f"BUILD SPEC — {c['meta'].get('ticker','')} {c['n_segments']}-SEGMENT SOTP", bold=True)
    items = c.get("buildspec") or _default_buildspec(c)
    for j, txt in enumerate(items):
        rr = 3 + j
        cell(ws, f"A{rr}", str(j), font=ST.FONT_NOTE, align="center")
        cell(ws, f"B{rr}", txt, font=ST.FONT_BASE, wrap=True, align="left")


def _default_buildspec(c):
    return [
        "PRIME DIRECTIVES: (1) Disclosed beats derived. (2) Derived needs two corroborations. "
        "(3) Every estimate is flagged yellow and explains itself. (4) Never deliver without verification.",
        "RESEARCH USED: latest 10-K/10-Q (balance sheet, segment table); most recent earnings release "
        "(segment revenue & adjusted operating profit); live peer prices.",
        f"SEGMENTS: {c['n_segments']} reportable segments — " + ", ".join(s["name"] for s in c["segments"]) + ".",
        "EBITDA BASIS: Segment Adj. EBITDA = segment adjusted operating profit + allocated depreciation. "
        "Acquired-intangible amortization stays at corporate.",
        "NET DEBT: Built from the latest balance sheet — ST debt + LT debt = gross; less cash = net debt. "
        "Watch for maturities reclassified from long-term to current (not a paydown).",
        "MULTIPLES: Peer-comp driven base case (Comps tab) — graded peers, regression, composite.",
        "VALUATION FLOW: Sum segment EVs − capitalized corporate cost − segment minority (fair value) − "
        "consolidated net debt − nonredeemable NCI − separation costs = equity value / FD shares.",
        "FLAGS: Yellow fill = estimate / refresh-critical. Live prices are sourced. Peer net debt & NTM "
        "EBITDA are best-estimates pending refresh from filings/consensus.",
        "VERIFICATION: Recalc to zero errors; independent recompute of per-share in all scenarios; "
        "entity split ties to Base; sensitivity base-case cell reproduces the headline; reverse SOTP sane.",
    ]


def build_spin(ws, c):
    spin = c["spin"]
    ST.setup_sheet(ws, ST.NOTE_GRAY, widths={"A": 30, "B": 118})
    cell(ws, "A1", f"{spin.get('entity_name','SPIN')} ({spin.get('entity_ticker','')}) SPIN-OFF — "
                   "STRUCTURE, TIMELINE & KEY FACTS", bold=True)
    if not spin.get("enabled"):
        cell(ws, "A3", "No spin-off configured for this company.", font=ST.FONT_NOTE)
        return
    facts = [
        ("Structure", spin.get("structure", "")),
        ("Timeline", spin.get("timeline", "")),
        ("Businesses", spin.get("businesses", "")),
        ("Financials", spin.get("financials", "")),
        ("Capitalization", spin.get("capitalization", "")),
        ("Proceeds", spin.get("proceeds", "")),
        ("Parent balance sheet", spin.get("parent_bs", "")),
        ("Leverage tie", spin.get("leverage_tie", "")),
        ("RemainCo", spin.get("remainco", "")),
        ("Guidance", spin.get("guidance", "")),
    ]
    r = 4
    for lab, txt in facts:
        if not txt:
            continue
        cell(ws, f"A{r}", lab, bold=True)
        cell(ws, f"B{r}", txt, font=ST.FONT_BASE, wrap=True, align="left")
        r += 1
    calls = c["meta"].get("judgment_calls", [])
    if calls:
        section(ws, r + 1, "KEY JUDGMENT CALLS", last_col="B")
        for j, txt in enumerate(calls):
            cell(ws, f"A{r+2+j}", f"- {txt}", font=ST.FONT_NOTE, wrap=True, align="left")
            ws.merge_cells(f"A{r+2+j}:B{r+2+j}")


def build_trends(ws, c):
    A = c["A"]
    tr = c["trends"]
    ST.setup_sheet(ws, ST.NOTE_GRAY, freeze="A4", widths={"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 66})
    cell(ws, "A1", f"RECENT TRENDS — {c['meta'].get('ticker','')} VS. COMPS", bold=True)
    cell(ws, "A2", "Compiled from company releases. Blue = reported.", font=ST.FONT_NOTE)
    quarters = tr.get("quarters", [])
    if not quarters:
        cell(ws, "A4", "Add quarterly revenue / margin / EPS history to the config to populate this tab.",
             font=ST.FONT_NOTE)
        return
    section(ws, 4, "QUARTERLY TREND", last_col=get_column_letter(1 + len(quarters)))
    cell(ws, "A5", "($M)")
    for j, q in enumerate(quarters):
        cell(ws, f"{get_column_letter(2+j)}5", q, bold=True, fill=ST.FILL_LIGHT, align="center")
    series = [("Revenue", tr.get("revenue", []), ST.FMT_USD_M),
              ("Operating margin (GAAP)", tr.get("op_margin", []), ST.FMT_PCT),
              ("Adj. diluted EPS ($/sh)", tr.get("adj_eps", []), ST.FMT_USD_2)]
    r = 6
    for lab, vals, fmt in series:
        cell(ws, f"A{r}", lab)
        for j, v in enumerate(vals):
            input_cell(ws, f"{get_column_letter(2+j)}{r}", v, fmt=fmt)
        r += 1
    seg_growth = tr.get("seg_growth", {})
    if seg_growth:
        section(ws, r + 1, "SEGMENT REVENUE GROWTH (latest y/y)", last_col="G")
        r += 2
        for name, g in seg_growth.items():
            cell(ws, f"A{r}", name)
            input_cell(ws, f"B{r}", g, fmt=ST.FMT_PCT)
            r += 1


def build_street(ws, c):
    A, S = c["A"], c["S"]
    st = c["street"]
    last_seg = c["seg_cols"][-1]
    min_idx = c["minority_idx"]
    min_col = c["seg_cols"][min_idx] if min_idx is not None else None
    ey = c["meta"].get("est_year", "2026E")
    ST.setup_sheet(ws, ST.NOTE_GRAY, freeze="A4", widths={"A": 26, "B": 16, "C": 12, "E": 14, "F": 16, "G": 50})
    cell(ws, "A1", "STREET COVERAGE — RATINGS, TARGETS & CONSENSUS", bold=True)
    cell(ws, "A2", "Compiled from public aggregators; estimates flagged. Refresh before use.", font=ST.FONT_NOTE)
    section(ws, 4, "CONSENSUS SNAPSHOT", last_col="C")
    cell(ws, "A5", "Number of covering analysts")
    input_cell(ws, "B5", st.get("analysts"), fmt=ST.FMT_NUM)
    cell(ws, "A6", "Consensus rating")
    input_cell(ws, "B6", st.get("rating", ""))
    cell(ws, "A7", "Consensus 12-mo price target ($/sh)")
    input_cell(ws, "B7", st.get("target"), fmt=ST.FMT_USD_2, estimate=True)
    note(ws, "G7", "Average across aggregators. Refresh.")
    cell(ws, "A8", "High / low target ($/sh)")
    input_cell(ws, "B8", st.get("high"), fmt=ST.FMT_USD_2)
    input_cell(ws, "C8", st.get("low"), fmt=ST.FMT_USD_2)
    cell(ws, "A9", "Consensus implied upside vs current")
    cell(ws, "B9", f"=B7/Assumptions!$B${A['price']}-1", fmt=ST.FMT_PCT)
    cell(ws, "A10", "Memo: model SOTP Base")
    cell(ws, "B10", f"=SOTP!C{S['per_share']}", fmt=ST.FMT_USD_2)

    section(ws, 12, "WHERE WE DIFFER — MODEL VS. STREET", last_col="D")
    for col, h in (("A", "Metric"), ("B", "Street/Cons"), ("C", "Model"), ("D", "Δ")):
        cell(ws, f"{col}13", h, bold=True, fill=ST.FILL_LIGHT)
    cell(ws, "G13", "Why", bold=True)
    cell(ws, "A14", "12-mo value / target ($/sh)")
    cell(ws, "B14", "=B7", fmt=ST.FMT_USD_2)
    cell(ws, "C14", f"=SOTP!C{S['per_share']}", fmt=ST.FMT_USD_2)
    cell(ws, "D14", "=C14/B14-1", fmt=ST.FMT_PCT)
    cell(ws, "A15", f"{ey} revenue ($M)")
    input_cell(ws, "B15", st.get("fy26_rev"), fmt=ST.FMT_USD_M, estimate=True)
    cell(ws, "C15", f"=SUM(SOTP!B{S['rev']}:{last_seg}{S['rev']})", fmt=ST.FMT_USD_M)
    cell(ws, "D15", "=C15/B15-1", fmt=ST.FMT_PCT)
    cell(ws, "A16", f"{ey} adj EBITDA ($M)")
    input_cell(ws, "B16", st.get("fy26_ebitda"), fmt=ST.FMT_USD_M, estimate=True)
    cell(ws, "C16", f"=SUM(SOTP!B{S['ebitda']}:{last_seg}{S['ebitda']})+Assumptions!B{A['e_corp']}", fmt=ST.FMT_USD_M)
    cell(ws, "D16", "=C16/B16-1", fmt=ST.FMT_PCT)

    section(ws, 18, "STREET-IMPLIED VALUATION", last_col="C")
    cell(ws, "A19", "Equity value at consensus PT")
    cell(ws, "B19", f"=B7*Assumptions!B{A['fd_shares']}", fmt=ST.FMT_USD_M)
    minterm = f"+Assumptions!B{A['minority_pct']}*SOTP!C{S['ev0']+min_idx}" if min_idx is not None else ""
    cell(ws, "A20", "Plus net debt+NCI+sep; less other assets; plus segment minority")
    cell(ws, "B20", f"=Assumptions!B{A['net_spin']}+Assumptions!B{A['nonredeem']}+Assumptions!B{A['sep_cost']}"
                    f"-Assumptions!B{A['retained']}-Assumptions!B{A['notes_rec']}{minterm}", fmt=ST.FMT_USD_M)
    cell(ws, "A21", "Street-implied operating EV")
    cell(ws, "B21", "=B19+B20", fmt=ST.FMT_USD_M, bold=True)
    cell(ws, "A22", f"Street-implied EV / {ey} EBITDA")
    cell(ws, "B22", f"=B21/(SUM(SOTP!B{S['ebitda']}:{last_seg}{S['ebitda']}))", fmt=ST.FMT_MULT)
    cell(ws, "A23", f"Memo: model Base combined EV / {ey} EBITDA")
    cell(ws, "B23", f"=SOTP!C{S['comb_ev']}/(SUM(SOTP!B{S['ebitda']}:{last_seg}{S['ebitda']}))", fmt=ST.FMT_MULT)

    catalysts = st.get("catalysts", [])
    if catalysts:
        section(ws, 25, "CATALYST CALENDAR", last_col="C")
        for j, (when, what) in enumerate(catalysts):
            rr = 26 + j
            cell(ws, f"A{rr}", when, bold=True)
            cell(ws, f"B{rr}", what, font=ST.FONT_BASE, wrap=True, align="left")
