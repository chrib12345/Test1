"""Formula-driven sheets: Assumptions, SOTP, Valuation Summary, Sensitivity,
Comps, FCF Build. All numeric output is computed by Excel formulas wired back
to the Assumptions sheet, exactly as in the SPGI master template.
"""
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from . import styles as ST
from .styles import cell, section, input_cell, note

SCEN = [("B", "Bear", "mult_bear"), ("C", "Base", "mult_base"), ("D", "Re-Rate", "mult_rerate")]


# --------------------------------------------------------------------------
# Assumptions — the input engine
# --------------------------------------------------------------------------
def build_assumptions(ws, c):
    A = c["A"]
    segs = c["segments"]
    m, nd, corp, fcf = c["market"], c["net_debt"], c["corporate"], c["fcf"]
    yr = c["meta"].get("fy_actual", "FY2025A")
    ey = c["meta"].get("est_year", "2026E")
    ST.setup_sheet(ws, ST.MED_BLUE, freeze="A4",
                   widths={"A": 40, "B": 13, "C": 13, "D": 12, "E": 12, "F": 12, "G": 58})

    cell(ws, "A1", f"{c['meta'].get('company','')} ({c['meta'].get('exchange','')}: "
                   f"{c['meta'].get('ticker','')}) — SOTP MODEL: ASSUMPTIONS & SCENARIO CONTROL", bold=True)
    cell(ws, "A2", "All $ in millions except per-share data.  Blue = inputs | Black = formulas | "
                   "Green = links | Yellow = estimate/refresh | Col G = source notes.", font=ST.FONT_NOTE)

    section(ws, A["mkt_hdr"], "MARKET DATA")
    r = A["price"]
    cell(ws, f"A{r}", "Current share price ($/sh)")
    cell(ws, f"B{r}", f"=IFERROR(IF(ISNUMBER($B${A['bb_tbl']+1}),$B${A['bb_tbl']+1},$D${r}),$D${r})",
         font=ST.FONT_LINK, fmt=ST.FMT_USD_2)
    cell(ws, f"C{r}", "live (feed) or manual", font=ST.FONT_NOTE)
    input_cell(ws, f"D{r}", m.get("price"), fmt=ST.FMT_USD_2, estimate=True)
    cell(ws, f"E{r}", "manual fallback", font=ST.FONT_NOTE)
    note(ws, f"G{r}", m.get("price_note", "Daily close. Stamp as-of date; refresh before delivery."))
    r = A["lowhigh"]
    cell(ws, f"A{r}", "52-week low / high ($/sh)")
    input_cell(ws, f"B{r}", m.get("low52"), fmt=ST.FMT_USD_2)
    input_cell(ws, f"C{r}", m.get("high52"), fmt=ST.FMT_USD_2)
    note(ws, f"G{r}", m.get("range_note", "Trailing 1-yr daily low/high."))
    r = A["basic_sh"]
    cell(ws, f"A{r}", "Basic shares outstanding (M)")
    input_cell(ws, f"B{r}", m.get("basic_shares"), fmt=ST.FMT_NUM1, estimate=m.get("shares_est", False))
    note(ws, f"G{r}", m.get("shares_note", "Latest filing, rolled forward for buybacks (est)."))
    r = A["dilutive"]
    cell(ws, f"A{r}", "Other dilutive securities (M)")
    input_cell(ws, f"B{r}", m.get("dilutive_shares", 0), fmt=ST.FMT_NUM1)
    note(ws, f"G{r}", "Diluted vs basic (10-Q).")
    r = A["mktcap"]
    cell(ws, f"A{r}", "Current market capitalization")
    cell(ws, f"B{r}", f"=B{A['price']}*(B{A['basic_sh']}+B{A['dilutive']})", fmt=ST.FMT_USD_M)

    # segment financials block (segments across columns)
    section(ws, A["seg_hdr"], f"SEGMENT FINANCIALS ({yr}) & {ey} DRIVERS")
    note(ws, f"G{A['seg_hdr']}", c["meta"].get("seg_source", "Source: latest annual segment disclosure"))
    cell(ws, f"A{A['seg_units']}", "($M)")
    for i, s in enumerate(segs):
        col = s["col"]
        cell(ws, f"{col}{A['seg_units']}", s["name"], bold=True, fill=ST.FILL_LIGHT, align="center")
    rows = [
        ("fy_rev", f"{yr} revenue", "fy_revenue", ST.FMT_USD_M, True),
        ("fy_op", f"{yr} adj. operating profit", "fy_op_profit", ST.FMT_USD_M, True),
        ("fy_dep", f"{yr} segment depreciation", "fy_deprec", ST.FMT_USD_M, True),
        ("growth", f"{ey} revenue growth (%)", "growth", ST.FMT_PCT, True),
        ("bps", f"{ey} adj. margin change (bps)", "margin_bps", ST.FMT_NUM, True),
    ]
    cell(ws, f"A{A['fy_opm']}", f"{yr} adj. operating margin")
    cell(ws, f"A{A['fy_ebitda']}", f"{yr} adj. EBITDA")
    cell(ws, f"A{A['fy_ebitdam']}", f"{yr} adj. EBITDA margin")
    for i, s in enumerate(segs):
        col = s["col"]
        input_cell(ws, f"{col}{A['fy_rev']}", s["fy_revenue"], fmt=ST.FMT_USD_M)
        input_cell(ws, f"{col}{A['fy_op']}", s["fy_op_profit"], fmt=ST.FMT_USD_M)
        cell(ws, f"{col}{A['fy_opm']}", f"={col}{A['fy_op']}/{col}{A['fy_rev']}", fmt=ST.FMT_PCT)
        input_cell(ws, f"{col}{A['fy_dep']}", s["fy_deprec"], fmt=ST.FMT_USD_M, estimate=True)
        cell(ws, f"{col}{A['fy_ebitda']}", f"={col}{A['fy_op']}+{col}{A['fy_dep']}", fmt=ST.FMT_USD_M)
        cell(ws, f"{col}{A['fy_ebitdam']}", f"={col}{A['fy_ebitda']}/{col}{A['fy_rev']}", fmt=ST.FMT_PCT)
        input_cell(ws, f"{col}{A['growth']}", s["growth"], fmt=ST.FMT_PCT, estimate=True)
        input_cell(ws, f"{col}{A['bps']}", s["margin_bps"], fmt=ST.FMT_NUM, estimate=True)
    for key, label, *_ in rows:
        cell(ws, f"A{A[key]}", label)
    note(ws, f"G{A['fy_dep']}", "Depreciation only (intangible amortization at corporate); allocated by revenue.")
    note(ws, f"G{A['growth']}", c["meta"].get("growth_note", "Management organic cc guidance."))
    note(ws, f"G{A['bps']}", "Conservative margin expansion on operating margin (est).")

    # corporate & minority
    section(ws, A["corp_hdr"], "CORPORATE / UNALLOCATED & MINORITY")
    cell(ws, f"A{A['fy_corp']}", f"{yr} corporate/unallocated EBITDA")
    input_cell(ws, f"B{A['fy_corp']}", corp.get("fy_corp_ebitda"), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['fy_corp']}", corp.get("corp_note", "Company-wide less segment sum."))
    cell(ws, f"A{A['e_corp']}", f"{ey} corporate/unallocated EBITDA")
    input_cell(ws, f"B{A['e_corp']}", corp.get("e_corp_ebitda"), fmt=ST.FMT_USD_M, estimate=True)
    note(ws, f"G{A['e_corp']}", "Held ~flat to prior year (est).")
    cell(ws, f"A{A['minority_pct']}", "Minority interest in a segment (%)")
    input_cell(ws, f"B{A['minority_pct']}", corp.get("minority_pct", 0), fmt=ST.FMT_PCT)
    note(ws, f"G{A['minority_pct']}", corp.get("minority_note",
         "Minority deducted at fair value (% x segment EV)."))

    # multiples block (segments across rows, scenarios across cols)
    section(ws, A["mult_hdr"], f"VALUATION MULTIPLES — EV / {ey} STANDALONE ADJ. EBITDA")
    note(ws, f"G{A['mult_hdr']}", "See Comps tab — graded peers, regression, grade-weighted composite")
    cell(ws, f"A{A['mult_units']}", "Segment", bold=True)
    for col, label, _ in SCEN:
        cell(ws, f"{col}{A['mult_units']}", label, bold=True, fill=ST.FILL_LIGHT, align="center")
    for i, s in enumerate(segs):
        rr = A["mult_first"] + i
        cell(ws, f"A{rr}", s["name"])
        input_cell(ws, f"B{rr}", s["mult_bear"], fmt=ST.FMT_MULT)
        input_cell(ws, f"C{rr}", s["mult_base"], fmt=ST.FMT_MULT)
        input_cell(ws, f"D{rr}", s["mult_rerate"], fmt=ST.FMT_MULT)
        if s.get("comp_anchor_note"):
            note(ws, f"G{rr}", s["comp_anchor_note"])

    _build_netdebt(ws, c)


def _build_netdebt(ws, c):
    A, nd, corp = c["A"], c["net_debt"], c["corporate"]
    spin = c["spin"]
    section(ws, A["nb_hdr"], "NET DEBT BUILD & RECONCILIATION ($M) — BUILD FROM BALANCE SHEET, NOT A RATIO")
    def lbl(key, text):
        cell(ws, f"A{A[key]}", text)
    lbl("bs_date", "Balance sheet date")
    input_cell(ws, f"B{A['bs_date']}", nd.get("bs_date", ""))
    cell(ws, f"C{A['bs_date']}", f"prior ({nd.get('prior_date','')})", font=ST.FONT_NOTE)
    note(ws, f"G{A['bs_date']}", "Every component below is same-dated.")
    lbl("st_debt", "Short-term debt / current maturities")
    input_cell(ws, f"B{A['st_debt']}", nd.get("st_debt"), fmt=ST.FMT_USD_M)
    input_cell(ws, f"C{A['st_debt']}", nd.get("prior_st"), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['st_debt']}", "Balance sheet, not derived.")
    lbl("lt_debt", "Long-term debt")
    input_cell(ws, f"B{A['lt_debt']}", nd.get("lt_debt"), fmt=ST.FMT_USD_M)
    input_cell(ws, f"C{A['lt_debt']}", nd.get("prior_lt"), fmt=ST.FMT_USD_M)
    lbl("other_debt", "Other debt-like adjustments")
    input_cell(ws, f"B{A['other_debt']}", nd.get("other_debt", 0), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['other_debt']}", "Leases out unless EBITDAR; NCI separate line.")
    lbl("gross_debt", "Gross debt")
    cell(ws, f"B{A['gross_debt']}", f"=SUM(B{A['st_debt']}:B{A['other_debt']})", fmt=ST.FMT_USD_M, bold=True)
    lbl("cash", "Less: cash & equivalents")
    input_cell(ws, f"B{A['cash']}", nd.get("cash"), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['cash']}", "Unrestricted only.")
    lbl("net_debt", "Net debt — at balance sheet date")
    cell(ws, f"B{A['net_debt']}", f"=B{A['gross_debt']}-B{A['cash']}", fmt=ST.FMT_USD_M, bold=True)
    lbl("stated_lev", "Company-stated gross leverage (x)")
    input_cell(ws, f"B{A['stated_lev']}", nd.get("stated_leverage"), fmt=ST.FMT_MULT)
    note(ws, f"G{A['stated_lev']}", "Stated gross leverage (incl. redeemable NCI as debt-like).")
    lbl("lev_ebitda", "Stated-leverage EBITDA basis")
    input_cell(ws, f"C{A['lev_ebitda']}", nd.get("leverage_ebitda_basis"), fmt=ST.FMT_USD_M)
    cell(ws, f"D{A['lev_ebitda']}", f"=B{A['stated_lev']}*C{A['lev_ebitda']}", fmt=ST.FMT_USD_M)
    note(ws, f"G{A['lev_ebitda']}", "D should approximate gross debt + redeemable NCI.")
    lbl("red_nci", "Redeemable NCI (minority of segment)")
    input_cell(ws, f"B{A['red_nci']}", nd.get("redeemable_nci", 0), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['red_nci']}", "Carrying value. Valued at FAIR value in SOTP — not added to net debt.")
    lbl("lev_tie", "Leverage tie check")
    cell(ws, f"B{A['lev_tie']}",
         f'=IF(ABS((B{A["gross_debt"]}+B{A["red_nci"]})/C{A["lev_ebitda"]}-B{A["stated_lev"]})<=0.1,'
         f'"TIES (gross incl NCI)","CHECK — reconcile")')
    lbl("buyback", "Subsequent: buybacks since BS date (+)")
    input_cell(ws, f"B{A['buyback']}", nd.get("buyback_rollfwd", 0), fmt=ST.FMT_USD_M, estimate=True)
    note(ws, f"G{A['buyback']}", "Repurchases after BS date ADD to net debt (est).")
    lbl("leak", "Subsequent: spin-notes net leakage/fees (+)")
    input_cell(ws, f"B{A['leak']}", nd.get("notes_leakage", 0), fmt=ST.FMT_USD_M, estimate=True)
    note(ws, f"G{A['leak']}", "Notes less proceeds to parent; fees/leakage (est).")
    lbl("net_spin", "NET DEBT — marked to spin date (feeds valuation)")
    cell(ws, f"B{A['net_spin']}", f"=B{A['net_debt']}+B{A['buyback']}+B{A['leak']}", fmt=ST.FMT_USD_M, bold=True)
    note(ws, f"G{A['net_spin']}", "THE cell the SOTP uses.")
    lbl("reclass", "Reclassification check")
    cell(ws, f"B{A['reclass']}",
         f'=IF((C{A["st_debt"]}+C{A["lt_debt"]})=0,"(enter prior ST/LT to enable)",'
         f'IF(AND((C{A["lt_debt"]}-B{A["lt_debt"]})>250,ABS((B{A["st_debt"]}+B{A["lt_debt"]})-'
         f'(C{A["st_debt"]}+C{A["lt_debt"]}))<250),"RECLASS: LT fell ~$"&TEXT(C{A["lt_debt"]}-B{A["lt_debt"]},'
         f'"#,##0")&"M but TOTAL borrowings ~flat — maturity moved to short-term, not repaid. '
         f'Use gross (ST+LT).","OK — change in gross borrowings is real"))')
    note(ws, f"G{A['reclass']}", "Prior-period ST/LT entered in col C to auto-detect reclassification.")
    lbl("nonredeem", "Nonredeemable NCI (equity)")
    input_cell(ws, f"B{A['nonredeem']}", nd.get("nonredeem_nci", 0), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['nonredeem']}", "Book value.")
    lbl("sep_cost", "Remaining one-time separation costs (after-tax)")
    input_cell(ws, f"B{A['sep_cost']}", nd.get("sep_costs", 0), fmt=ST.FMT_USD_M, estimate=True)
    note(ws, f"G{A['sep_cost']}", "Residual spin costs through completion (est).")

    section(ws, A["oa_hdr"], "OTHER ASSETS (retained stakes / notes — default 0)")
    cell(ws, f"A{A['retained']}", "Retained stakes / JV interests (value)")
    input_cell(ws, f"B{A['retained']}", nd.get("retained_stakes", 0), fmt=ST.FMT_USD_M)
    note(ws, f"G{A['retained']}", "Asset lines in the bridge — never netted into debt.")
    cell(ws, f"A{A['notes_rec']}", "Notes receivable / other assets")
    input_cell(ws, f"B{A['notes_rec']}", nd.get("notes_receivable", 0), fmt=ST.FMT_USD_M)

    section(ws, A["ed_hdr"], "POST-SPIN ENTITY DEBT ALLOCATION ($M)")
    cell(ws, f"A{A['spin_nd']}", f"{spin.get('entity_ticker','Spin')} net debt at spin")
    input_cell(ws, f"B{A['spin_nd']}", spin.get("net_debt", 0), fmt=ST.FMT_USD_M, estimate=True)
    note(ws, f"G{A['spin_nd']}", spin.get("net_debt_note", "Spin notes less retained cash."))
    cell(ws, f"A{A['remainco_nd']}", "RemainCo pro forma net debt")
    cell(ws, f"B{A['remainco_nd']}", f"=B{A['net_spin']}-B{A['spin_nd']}", fmt=ST.FMT_USD_M)
    note(ws, f"G{A['remainco_nd']}", "Consolidated less spin net debt that departs.")

    section(ws, A["sc_hdr"], "SHARE COUNT")
    cell(ws, f"A{A['fd_shares']}", "Fully diluted shares (M)")
    cell(ws, f"B{A['fd_shares']}", f"=B{A['basic_sh']}+B{A['dilutive']}", fmt=ST.FMT_NUM1, bold=True)
    note(ws, f"G{A['fd_shares']}", "Cross-check vs guidance weighted-average.")

    # optional live-feed scaffold (inert)
    section(ws, A["bb_hdr"], "LIVE FEED (OPTIONAL) — REQUIRES MARKET-DATA EXCEL ADD-IN")
    tk = c["meta"].get("ticker", "TICKER")
    cell(ws, f"A{A['bb_intro']}", f"To activate: in B{A['bb_tbl']+1} replace with a live formula "
         f'(e.g. =BDP("{tk} US Equity","PX_LAST")). B{A["price"]} auto-detects and switches to live.',
         font=ST.FONT_NOTE)
    cell(ws, f"A{A['bb_tbl']}", "Field", bold=True)
    cell(ws, f"B{A['bb_tbl']}", "Formula (inert — prepend = to activate)", bold=True)
    feeds = [("Last price", "PX_LAST"), ("52-week low", "LOW_52WEEK"),
             ("52-week high", "HIGH_52WEEK"), ("Shares outstanding (M)", "EQY_SH_OUT"),
             ("Consensus 12-mo target", "BEST_TARGET_PRICE")]
    for j, (fld, code) in enumerate(feeds):
        rr = A["bb_tbl"] + 1 + j
        cell(ws, f"A{rr}", fld)
        cell(ws, f"B{rr}", f'BDP("{tk} US Equity","{code}")', font=ST.FONT_NOTE)


# --------------------------------------------------------------------------
# SOTP — the valuation engine
# --------------------------------------------------------------------------
def build_sotp(ws, c):
    A, S = c["A"], c["S"]
    segs = c["segments"]
    last_seg = c["seg_cols"][-1]
    ey = c["meta"].get("est_year", "2026E")
    ST.setup_sheet(ws, ST.NAVY, freeze="A4", widths={"A": 42, "B": 13, "C": 13, "D": 13, "F": 40})

    cell(ws, "A1", f"SUM-OF-THE-PARTS — EV / {ey} STANDALONE ADJ. EBITDA ({c['n_segments']} SEGMENTS)", bold=True)
    cell(ws, "A2", "$ in millions except per-share data. Engine is formula-driven from Assumptions.", font=ST.FONT_NOTE)

    section(ws, S["build_hdr"], f"{ey} STANDALONE ADJ. EBITDA BUILD", last_col=last_seg)
    cell(ws, f"A{S['seg']}", "Segment", bold=True)
    for s in segs:
        cell(ws, f"{s['col']}{S['seg']}", s["short"], bold=True, fill=ST.FILL_LIGHT, align="center")
    cell(ws, f"A{S['rev']}", f"{ey} revenue")
    cell(ws, f"A{S['opm']}", f"{ey} adj. operating margin")
    cell(ws, f"A{S['op']}", f"{ey} adj. operating profit")
    cell(ws, f"A{S['dep']}", f"{ey} depreciation")
    cell(ws, f"A{S['ebitda']}", f"{ey} adj. EBITDA")
    cell(ws, f"A{S['ebitdam']}", f"{ey} adj. EBITDA margin")
    for s in segs:
        col = s["col"]
        cell(ws, f"{col}{S['rev']}", f"=Assumptions!{col}{A['fy_rev']}*(1+Assumptions!{col}{A['growth']})", fmt=ST.FMT_USD_M)
        cell(ws, f"{col}{S['opm']}", f"=Assumptions!{col}{A['fy_opm']}+Assumptions!{col}{A['bps']}/10000", fmt=ST.FMT_PCT)
        cell(ws, f"{col}{S['op']}", f"={col}{S['rev']}*{col}{S['opm']}", fmt=ST.FMT_USD_M)
        cell(ws, f"{col}{S['dep']}", f"=Assumptions!{col}{A['fy_dep']}*(1+Assumptions!{col}{A['growth']})", fmt=ST.FMT_USD_M)
        cell(ws, f"{col}{S['ebitda']}", f"={col}{S['op']}+{col}{S['dep']}", fmt=ST.FMT_USD_M, bold=True)
        cell(ws, f"{col}{S['ebitdam']}", f"={col}{S['ebitda']}/{col}{S['rev']}", fmt=ST.FMT_PCT)

    # valuation bridge
    section(ws, S["bridge_hdr"], "SOTP VALUATION BRIDGE", last_col="D")
    for col, label, _ in SCEN:
        cell(ws, f"{col}{S['scen']}", label, bold=True, fill=ST.FILL_LIGHT, align="center")
    min_idx = c["minority_idx"]
    for i, s in enumerate(segs):
        rr = S["ev0"] + i
        suffix = " (100%)" if i == min_idx else ""
        cell(ws, f"A{rr}", f"{s['name']} EV{suffix}")
        for col, _, _ in SCEN:
            mult_row = A["mult_first"] + i
            cell(ws, f"{col}{rr}", f"=${s['col']}${S['ebitda']}*Assumptions!${col}${mult_row}", fmt=ST.FMT_USD_M)

    def line(key, label, formula_for, fmt=ST.FMT_USD_M, bold=False, cols="BCD", notecol=None):
        cell(ws, f"A{S[key]}", label, bold=bold)
        for col in cols:
            cell(ws, f"{col}{S[key]}", formula_for(col), fmt=fmt, bold=bold)
        if notecol:
            note(ws, f"F{S[key]}", notecol)

    ev0, evl = S["ev0"], S["ev_last"]
    line("sum", "Sum of segment enterprise values",
         lambda col: f"=SUM({col}{ev0}:{col}{evl})", bold=True)
    line("corp", "Less: capitalized corporate/unallocated cost",
         lambda col: f"=Assumptions!$B${A['e_corp']}*({col}{S['sum']}/SUM($B${S['ebitda']}:${last_seg}${S['ebitda']}))",
         notecol="corp EBITDA x blended company multiple")
    if min_idx is not None:
        minrow = S["ev0"] + min_idx
        line("minority", "Less: segment minority (fair value)",
             lambda col: f"=-Assumptions!$B${A['minority_pct']}*{col}{minrow}")
    else:
        line("minority", "Less: segment minority (fair value)", lambda col: "=0")
    line("comb_ev", "Combined operating enterprise value",
         lambda col: f"={col}{S['sum']}+{col}{S['corp']}+{col}{S['minority']}", bold=True)
    line("retained", "Plus: retained stakes / JV interests", lambda col: f"=Assumptions!$B${A['retained']}")
    line("notes", "Plus: notes receivable / other assets", lambda col: f"=Assumptions!$B${A['notes_rec']}")
    line("net_debt", "Less: consolidated net debt", lambda col: f"=-Assumptions!$B${A['net_spin']}")
    line("nonredeem", "Less: nonredeemable NCI", lambda col: f"=-Assumptions!$B${A['nonredeem']}")
    line("sep_cost", "Less: remaining separation costs", lambda col: f"=-Assumptions!$B${A['sep_cost']}")
    line("equity", "Implied equity value",
         lambda col: f"=SUM({col}{S['comb_ev']}:{col}{S['sep_cost']})", bold=True)
    line("shares", "Fully diluted shares (M)", lambda col: f"=Assumptions!$B${A['fd_shares']}", fmt=ST.FMT_NUM1)
    line("per_share", "Implied value per share",
         lambda col: f"={col}{S['equity']}/{col}{S['shares']}", fmt=ST.FMT_USD_2, bold=True)
    line("cur_price", "Current share price", lambda col: f"=Assumptions!$B${A['price']}", fmt=ST.FMT_USD_2)
    line("upside", "Implied upside / (downside)",
         lambda col: f"={col}{S['per_share']}/{col}{S['cur_price']}-1", fmt=ST.FMT_PCT)

    if c["spin"].get("enabled") and c["spin_idx"] is not None:
        _build_entity_split(ws, c)
    _build_reverse(ws, c)


def _build_entity_split(ws, c):
    A, S = c["A"], c["S"]
    spin = c["spin"]
    spin_ev_row = S["ev0"] + c["spin_idx"]
    en = spin.get("entity_name", "Spin entity")
    tk = spin.get("entity_ticker", "SPIN")
    section(ws, S["ent_hdr"], f"IMPLIED VALUE BY ENTITY — BASE CASE (RemainCo vs {tk})", last_col="D")
    cell(ws, f"A{S['remainco']}", "RemainCo (ex-spin) equity value")
    cell(ws, f"B{S['remainco']}", f"=C{S['equity']}-B{S['spin_eq']}", fmt=ST.FMT_USD_M)
    note(ws, f"F{S['remainco']}", "Total Base equity less spin equity")
    cell(ws, f"A{S['spin_eq']}", f"{en} ({tk}) equity value")
    cell(ws, f"B{S['spin_eq']}", f"=C{spin_ev_row}-Assumptions!B{A['spin_nd']}", fmt=ST.FMT_USD_M)
    note(ws, f"F{S['spin_eq']}", "Spin Base EV less net debt at spin")
    cell(ws, f"A{S['remain_ps']}", "RemainCo value per current share")
    cell(ws, f"B{S['remain_ps']}", f"=B{S['remainco']}/Assumptions!B{A['fd_shares']}", fmt=ST.FMT_USD_2)
    cell(ws, f"A{S['spin_ps']}", f"{en} ({tk}) per current share")
    cell(ws, f"B{S['spin_ps']}", f"=B{S['spin_eq']}/Assumptions!B{A['fd_shares']}", fmt=ST.FMT_USD_2)
    cell(ws, f"A{S['ent_total']}", "Total (ties to Base case above)", bold=True)
    cell(ws, f"B{S['ent_total']}", f"=B{S['remain_ps']}+B{S['spin_ps']}", fmt=ST.FMT_USD_2, bold=True)
    note(ws, f"F{S['ent_total']}", f"Must equal SOTP!C{S['per_share']}")


def _build_reverse(ws, c):
    A, S = c["A"], c["S"]
    segs = c["segments"]
    last_seg = c["seg_cols"][-1]
    min_idx, row_idx = c["minority_idx"], c["sens_row_idx"]
    min_col = c["seg_cols"][min_idx] if min_idx is not None else None
    row_seg = segs[row_idx]
    section(ws, S["rev_hdr"], "MARKET-IMPLIED — REVERSE SOTP AT CURRENT PRICE", last_col="D")
    cell(ws, f"A{S['mkt_eq']}", "Market equity value (FD basis)")
    cell(ws, f"B{S['mkt_eq']}", f"=Assumptions!B{A['price']}*Assumptions!B{A['fd_shares']}", fmt=ST.FMT_USD_M)
    cell(ws, f"A{S['plus_nd']}", "Plus net debt + nonredeem NCI + sep costs; less other assets")
    cell(ws, f"B{S['plus_nd']}",
         f"=Assumptions!B{A['net_spin']}+Assumptions!B{A['nonredeem']}+Assumptions!B{A['sep_cost']}"
         f"-Assumptions!B{A['retained']}-Assumptions!B{A['notes_rec']}", fmt=ST.FMT_USD_M)
    cell(ws, f"A{S['mkt_ev']}", "Market-implied attributable operating EV")
    cell(ws, f"B{S['mkt_ev']}", f"=B{S['mkt_eq']}+B{S['plus_nd']}", fmt=ST.FMT_USD_M, bold=True)
    minus = f"-Assumptions!B{A['minority_pct']}*{min_col}{S['ebitda']}" if min_col else ""
    cell(ws, f"A{S['blended']}", "Implied blended EV / attributable EBITDA")
    cell(ws, f"B{S['blended']}",
         f"=B{S['mkt_ev']}/(SUM(B{S['ebitda']}:{last_seg}{S['ebitda']}){minus})", fmt=ST.FMT_MULT)
    cell(ws, f"A{S['memo']}", "Memo: Base-case combined operating EV")
    cell(ws, f"B{S['memo']}", f"=C{S['comb_ev']}", fmt=ST.FMT_USD_M)
    # held-EV sum for the reverse solve (all but the solved row segment)
    terms = []
    for i, s in enumerate(segs):
        if i == row_idx:
            continue
        evrow = S["ev0"] + i
        if i == min_idx:
            terms.append(f"C{evrow}*(1-Assumptions!B{A['minority_pct']})")
        else:
            terms.append(f"C{evrow}")
    held = "+".join(terms)
    cell(ws, f"A{S['impl_row']}", f"Implied {row_seg['short']} EV/EBITDA (others at Base)")
    cell(ws, f"B{S['impl_row']}",
         f"=(B{S['mkt_ev']}-({held})-C{S['corp']})/{row_seg['col']}{S['ebitda']}", fmt=ST.FMT_MULT)
    note(ws, f"F{S['impl_row']}", "Market-implied multiple for the franchise segment vs its comps.")
    cell(ws, f"A{S['vs_base']}", f"  vs. Base-case {row_seg['short']} multiple")
    cell(ws, f"B{S['vs_base']}", f"=Assumptions!C{A['mult_first']+row_idx}", fmt=ST.FMT_MULT)
    # anchor peer = first comp of the row segment
    grp = c["CMP"]["groups"][row_idx]
    if grp["k"] > 0:
        anchor = segs[row_idx]["comps"][0].get("company", "anchor peer")
        cell(ws, f"A{S['vs_peer']}", f"  vs. {anchor} current EV/EBITDA")
        cell(ws, f"B{S['vs_peer']}", f"=Comps!K{grp['first']}", fmt=ST.FMT_MULT)


# --------------------------------------------------------------------------
# Valuation Summary
# --------------------------------------------------------------------------
def build_valuation_summary(ws, c):
    A, S = c["A"], c["S"]
    row_idx, col_idx = c["sens_row_idx"], c["sens_col_idx"]
    ey = c["meta"].get("est_year", "2026E")
    ST.setup_sheet(ws, ST.NAVY, widths={"A": 46, "B": 13, "C": 13, "D": 13, "G": 28, "H": 11, "I": 11, "J": 11, "K": 11})
    cell(ws, "A1", f"{c['meta'].get('company','')} — VALUATION SUMMARY", bold=True)
    cell(ws, "A2", f"Per-share values; see SOTP and Sensitivity tabs for detail.", font=ST.FONT_NOTE)
    section(ws, 4, "SCENARIO SUMMARY", last_col="D")
    for col, label, _ in SCEN:
        cell(ws, f"{col}5", label, bold=True, fill=ST.FILL_LIGHT, align="center")
    cell(ws, "A6", f"{c['segments'][row_idx]['short']} EV / {ey} EBITDA")
    cell(ws, "A7", f"{c['segments'][col_idx]['short']} EV / {ey} EBITDA")
    for col in "BCD":
        cell(ws, f"{col}6", f"=Assumptions!{col}{A['mult_first']+row_idx}", fmt=ST.FMT_MULT)
        cell(ws, f"{col}7", f"=Assumptions!{col}{A['mult_first']+col_idx}", fmt=ST.FMT_MULT)
        cell(ws, f"{col}8", f"=SOTP!{col}{S['per_share']}", fmt=ST.FMT_USD_2, bold=True)
        cell(ws, f"{col}9", f"=SOTP!{col}{S['upside']}", fmt=ST.FMT_PCT)
    cell(ws, "A8", "Implied value per share", bold=True)
    cell(ws, "A9", "Upside / (downside)")
    cell(ws, "A11", "Current share price")
    cell(ws, "B11", f"=Assumptions!B{A['price']}", fmt=ST.FMT_USD_2)
    if c["spin"].get("enabled"):
        cell(ws, "A12", "Base case — RemainCo / spin per share")
        cell(ws, "B12", f"=SOTP!B{S['remain_ps']}", fmt=ST.FMT_USD_2)
        cell(ws, "C12", f"=SOTP!B{S['spin_ps']}", fmt=ST.FMT_USD_2)
    cell(ws, "A13", "Consensus 12-mo target (Street)")
    cell(ws, "B13", "='Street Coverage'!B7", fmt=ST.FMT_USD_2)

    # football field
    section(ws, 15, "FOOTBALL FIELD — IMPLIED SHARE-PRICE RANGES", last_col="D")
    cell(ws, "G15", "chart data", font=ST.FONT_NOTE)
    for col, h in (("G", "Label"), ("H", "Low"), ("I", "High"), ("J", "Base"), ("K", "Range")):
        cell(ws, f"{col}16", h, bold=True, fill=ST.FILL_LIGHT, align="center")
    rows = [
        ("52-Week Trading Range", f"=Assumptions!B{A['lowhigh']}", f"=Assumptions!C{A['lowhigh']}"),
        ("Sensitivity Grid (Min-Max)", "=MIN(Sensitivity!B6:J14)", "=MAX(Sensitivity!B6:J14)"),
        ("SOTP Scenarios (Bear-ReRate)", f"=SOTP!B{S['per_share']}", f"=SOTP!D{S['per_share']}"),
        ("Street target range", "='Street Coverage'!C8", "='Street Coverage'!B8"),
        ("Current Price", f"=Assumptions!B{A['price']}-0.5", f"=Assumptions!B{A['price']}+0.5"),
    ]
    for j, (label, lo, hi) in enumerate(rows):
        rr = 17 + j
        cell(ws, f"G{rr}", label)
        cell(ws, f"H{rr}", lo, fmt=ST.FMT_USD_2)
        cell(ws, f"I{rr}", hi, fmt=ST.FMT_USD_2)
        cell(ws, f"J{rr}", f"=H{rr}", fmt=ST.FMT_USD_2)
        cell(ws, f"K{rr}", f"=I{rr}-H{rr}", fmt=ST.FMT_USD_2)


# --------------------------------------------------------------------------
# Sensitivity — two heat-mapped grids
# --------------------------------------------------------------------------
def build_sensitivity(ws, c):
    A, S = c["A"], c["S"]
    segs = c["segments"]
    sens = c["sensitivity"]
    row_idx, col_idx, min_idx = c["sens_row_idx"], c["sens_col_idx"], c["minority_idx"]
    row_steps, col_steps = sens["row_steps"], sens["col_steps"]
    ST.setup_sheet(ws, ST.NAVY, widths={"A": 16, "B": 11, "C": 11, "D": 11, "E": 11,
                                        "F": 11, "G": 11, "H": 11, "I": 11, "J": 11})
    rs, cs = segs[row_idx]["short"], segs[col_idx]["short"]
    cell(ws, "A1", "SENSITIVITY — IMPLIED SHARE PRICE", bold=True)
    cell(ws, "A2", f"{rs} EV/EBITDA (rows) vs {cs} EV/EBITDA (cols). Other segments held at Base; "
                   "capital structure constant.", font=ST.FONT_NOTE)

    grid_cols = [get_column_letter(2 + k) for k in range(9)]   # B..J

    def equity_formula(rcell, ccell):
        """Equity-from-scratch with row/col segments on variable multiples."""
        terms = []
        for i, s in enumerate(segs):
            evrow = S["ev0"] + i
            mfac = f"*(1-Assumptions!$B${A['minority_pct']})" if i == min_idx else ""
            if i == row_idx:
                terms.append(f"SOTP!${segs[row_idx]['col']}${S['ebitda']}*{rcell}{mfac}")
            elif i == col_idx:
                terms.append(f"SOTP!${segs[col_idx]['col']}${S['ebitda']}*{ccell}{mfac}")
            else:
                terms.append(f"SOTP!$C${evrow}{mfac}")
        body = "+".join(terms)
        body += f"+SOTP!$C${S['corp']}"
        body += (f"+Assumptions!$B${A['retained']}+Assumptions!$B${A['notes_rec']}"
                 f"-Assumptions!$B${A['net_spin']}-Assumptions!$B${A['nonredeem']}"
                 f"-Assumptions!$B${A['sep_cost']}")
        return f"=({body})/Assumptions!$B${A['fd_shares']}"

    # ---- grid 1: implied price ----
    cell(ws, "A4", "Implied share price ($/sh)", bold=True)
    cell(ws, "A5", f"{rs[:3]} x \\ {cs[:3]} x", bold=True, fill=ST.FILL_LIGHT)
    for k, gc in enumerate(grid_cols):
        cell(ws, f"{gc}5", col_steps[k], bold=True, fill=ST.FILL_LIGHT, align="center", fmt=ST.FMT_MULT)
    for ri in range(9):
        rr = 6 + ri
        cell(ws, f"A{rr}", row_steps[ri], bold=True, fill=ST.FILL_LIGHT, fmt=ST.FMT_MULT)
        for k, gc in enumerate(grid_cols):
            f = equity_formula(f"$A{rr}", f"{gc}$5")
            cell(ws, f"{gc}{rr}", f, fmt=ST.FMT_USD_2)
    ws.conditional_formatting.add("B6:J14", ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))

    # ---- grid 2: upside ----
    cell(ws, "A16", "Implied upside / (downside) vs current price", bold=True)
    cell(ws, "A17", f"{rs[:3]} x \\ {cs[:3]} x", bold=True, fill=ST.FILL_LIGHT)
    for k, gc in enumerate(grid_cols):
        cell(ws, f"{gc}17", col_steps[k], bold=True, fill=ST.FILL_LIGHT, align="center", fmt=ST.FMT_MULT)
    for ri in range(9):
        rr = 18 + ri
        src = 6 + ri
        cell(ws, f"A{rr}", row_steps[ri], bold=True, fill=ST.FILL_LIGHT, fmt=ST.FMT_MULT)
        for gc in grid_cols:
            cell(ws, f"{gc}{rr}", f"={gc}{src}/Assumptions!$B${A['price']}-1", fmt=ST.FMT_PCT)
    ws.conditional_formatting.add("B18:J26", ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))
    cell(ws, "A28", f"Heat map: red = low, green = high. Base case = "
         f"{rs} {segs[row_idx]['mult_base']}x / {cs} {segs[col_idx]['mult_base']}x.", font=ST.FONT_NOTE)


# --------------------------------------------------------------------------
# Comps — graded peers, live EV build, regression
# --------------------------------------------------------------------------
def build_comps(ws, c):
    A, S = c["A"], c["S"]
    segs = c["segments"]
    CMP = c["CMP"]
    ST.setup_sheet(ws, ST.NAVY, freeze="B4", widths={"A": 24, "B": 9, "C": 7, "D": 6, "E": 10,
                   "F": 9, "G": 11, "H": 12, "I": 11, "J": 11, "K": 11, "L": 12, "M": 30})
    cell(ws, "A1", "COMPS — GRADED PEERS, LIVE EV BUILD & MULTIPLE JUSTIFICATION", bold=True)
    cell(ws, "A2", "Prices live; net debt & NTM EBITDA are best-estimates pending refresh (yellow). "
                   "EV=Mkt cap+Net debt. Grade A=anchor (wt 2.0), B=context (1.0), C=loose (0.5).", font=ST.FONT_NOTE)
    hdrs = ["Company", "Ticker", "Grade", "Wt", "Price ($)", "Shares (M)", "Net debt",
            "NTM EBITDA", "Mkt cap", "EV", "EV/EBITDA", "EBITDA mgn", "As-of / note"]
    for j, h in enumerate(hdrs):
        cell(ws, f"{get_column_letter(1+j)}{CMP['header']}", h, bold=True, fill=ST.FILL_NAVY, font=ST.FONT_HDR, align="center")

    for i, s in enumerate(segs):
        grp = CMP["groups"][i]
        section(ws, grp["title"], f"{s['name'].upper()} — peers ({s.get('comp_group_title','')})", last_col="M")
        for j, p in enumerate(s.get("comps", [])):
            rr = grp["first"] + j
            cell(ws, f"A{rr}", p.get("company", ""))
            cell(ws, f"B{rr}", p.get("ticker", ""))
            cell(ws, f"C{rr}", p.get("grade", "B"), align="center")
            cell(ws, f"D{rr}", p.get("weight", 1), align="center", fmt=ST.FMT_NUM1)
            input_cell(ws, f"E{rr}", p.get("price"), fmt=ST.FMT_USD_2)
            input_cell(ws, f"F{rr}", p.get("shares"), fmt=ST.FMT_NUM)
            input_cell(ws, f"G{rr}", p.get("net_debt"), fmt=ST.FMT_USD_M, estimate=True)
            input_cell(ws, f"H{rr}", p.get("ntm_ebitda"), fmt=ST.FMT_USD_M, estimate=True)
            cell(ws, f"I{rr}", f"=E{rr}*F{rr}", fmt=ST.FMT_USD_M)
            cell(ws, f"J{rr}", f"=I{rr}+G{rr}", fmt=ST.FMT_USD_M)
            cell(ws, f"K{rr}", f"=J{rr}/H{rr}", fmt=ST.FMT_MULT)
            input_cell(ws, f"L{rr}", p.get("ebitda_margin"), fmt=ST.FMT_PCT)
            cell(ws, f"M{rr}", p.get("note", ""), font=ST.FONT_NOTE)
        # subtotal row
        sr = grp["subtotal"]
        cell(ws, f"A{sr}", "  Median / wtd-avg / selected base", font=ST.FONT_NOTE)
        if grp["k"] > 0:
            f, l = grp["first"], grp["last"]
            num = "+".join([f"D{x}*K{x}" for x in range(f, l + 1)])
            den = "+".join([f"D{x}" for x in range(f, l + 1)])
            cell(ws, f"G{sr}", "wtd ->", font=ST.FONT_NOTE)
            cell(ws, f"H{sr}", f"=({num})/({den})", fmt=ST.FMT_MULT)
            cell(ws, f"I{sr}", "median ->", font=ST.FONT_NOTE)
            cell(ws, f"J{sr}", f"=MEDIAN(K{f}:K{l})", fmt=ST.FMT_MULT)
        cell(ws, f"K{sr}", f"=Assumptions!C{A['mult_first']+i}", font=ST.FONT_LINK, fmt=ST.FMT_MULT)
        cell(ws, f"L{sr}", "selected ^", font=ST.FONT_NOTE)
        cell(ws, f"M{sr}", "H=grade-wtd avg, J=median, K=selected base (link)", font=ST.FONT_NOTE)

    _build_regression(ws, c)


def _build_regression(ws, c):
    A, S = c["A"], c["S"]
    segs = c["segments"]
    CMP = c["CMP"]
    # unique peers across all segment groups
    peers = []
    seen = set()
    for i, s in enumerate(segs):
        grp = CMP["groups"][i]
        for j, p in enumerate(s.get("comps", [])):
            tk = p.get("ticker", "")
            if tk and tk not in seen:
                seen.add(tk)
                peers.append((tk, grp["first"] + j))
    hdr = CMP["reg_hdr"]
    section(ws, hdr, "REGRESSION — EV/EBITDA vs EBITDA MARGIN (all unique peers)", last_col="M")
    cell(ws, f"A{hdr+1}", "Peer", bold=True)
    cell(ws, f"B{hdr+1}", "EBITDA mgn (x)", bold=True)
    cell(ws, f"C{hdr+1}", "EV/EBITDA (y)", bold=True)
    cell(ws, f"E{hdr+1}", "Regression diagnostics", bold=True)
    first = CMP["reg_first"]
    last = first + len(peers) - 1
    for k, (tk, srow) in enumerate(peers):
        rr = first + k
        cell(ws, f"A{rr}", tk)
        cell(ws, f"B{rr}", f"=L{srow}", fmt=ST.FMT_PCT)
        cell(ws, f"C{rr}", f"=K{srow}", fmt=ST.FMT_MULT)
    if peers:
        diag = [("Slope (per 1.00 margin)", f"=SLOPE(C{first}:C{last},B{first}:B{last})"),
                ("Intercept", f"=INTERCEPT(C{first}:C{last},B{first}:B{last})"),
                ("R-squared", f"=RSQ(C{first}:C{last},B{first}:B{last})"),
                ("n", f"=COUNT(C{first}:C{last})")]
        for k, (lab, f) in enumerate(diag):
            rr = first + k
            cell(ws, f"E{rr}", lab)
            cell(ws, f"G{rr}", f, fmt=ST.FMT_NUM1)
        c["CMP"]["slope_cell"] = f"$G${first}"
        c["CMP"]["intercept_cell"] = f"$G${first+1}"
        # regression-implied multiple per segment
        base = first + 6
        cell(ws, f"E{base-1}", "Regression-implied multiple at segment margin:", font=ST.FONT_NOTE)
        for i, s in enumerate(segs):
            rr = base + i
            cell(ws, f"E{rr}", s["name"])
            cell(ws, f"F{rr}", f"=SOTP!{s['col']}{S['ebitdam']}", fmt=ST.FMT_PCT)
            cell(ws, f"G{rr}", f"={c['CMP']['intercept_cell']}+{c['CMP']['slope_cell']}*SOTP!{s['col']}{S['ebitdam']}", fmt=ST.FMT_MULT)
            cell(ws, f"H{rr}", f"=Assumptions!C{A['mult_first']+i}", font=ST.FONT_LINK, fmt=ST.FMT_MULT)
            cell(ws, f"I{rr}", "selected base", font=ST.FONT_NOTE)


# --------------------------------------------------------------------------
# FCF Build
# --------------------------------------------------------------------------
def build_fcf(ws, c):
    A, S = c["A"], c["S"]
    fcf = c["fcf"]
    last_seg = c["seg_cols"][-1]
    ey = c["meta"].get("est_year", "2026E")
    y2 = c["meta"].get("est_year2", "2027E")
    ST.setup_sheet(ws, ST.NAVY, freeze="A4", widths={"A": 46, "B": 12, "C": 12, "D": 3, "G": 62})
    cell(ws, "A1", "FREE CASH FLOW BUILD — CONSOLIDATED", bold=True)
    cell(ws, "A2", "$ in millions except per-share data. Pre-spin consolidated basis.", font=ST.FONT_NOTE)
    section(ws, 4, "PRIOR-YEAR CASH FLOW ANCHOR", last_col="C")
    cell(ws, "A5", "Adjusted free cash flow — prior FY")
    input_cell(ws, "B5", fcf.get("fy_fcf"), fmt=ST.FMT_USD_M)
    note(ws, "G5", fcf.get("fcf_note", "Capital return implies adj FCF."))
    cell(ws, "A6", "Capital returned (buybacks + dividends)")
    input_cell(ws, "B6", fcf.get("capital_returned"), fmt=ST.FMT_USD_M)
    anchor = fcf.get("fy_ebitda_anchor", 1) or 1
    cell(ws, "A7", "Prior-FY adj FCF conversion (% of adj EBITDA)")
    cell(ws, "B7", f"=B5/{anchor}", fmt=ST.FMT_PCT)
    note(ws, "G7", "Prior FY adj EBITDA anchor.")

    section(ws, 9, "FCF BUILD — FORECAST", last_col="C")
    cell(ws, "B10", ey, bold=True, fill=ST.FILL_LIGHT, align="center")
    cell(ws, "C10", y2, bold=True, fill=ST.FILL_LIGHT, align="center")
    g = 1 + (fcf.get("y2_growth", 0.07))
    cell(ws, "A11", "Adjusted EBITDA (consolidated)")
    cell(ws, "B11", f"=SUM(SOTP!B{S['ebitda']}:{last_seg}{S['ebitda']})+Assumptions!B{A['e_corp']}", fmt=ST.FMT_USD_M)
    cell(ws, "C11", f"=B11*{g}", fmt=ST.FMT_USD_M)
    note(ws, "G11", f"{ey} = sum of segment EBITDA + corporate; {y2} growth applied.")
    cell(ws, "A12", "Less: cash interest")
    input_cell(ws, "B12", fcf.get("e_interest"), fmt=ST.FMT_USD_M)
    input_cell(ws, "C12", fcf.get("e_interest_y2", fcf.get("e_interest")), fmt=ST.FMT_USD_M)
    cell(ws, "A13", "Cash tax rate")
    input_cell(ws, "B13", fcf.get("tax_rate"), fmt=ST.FMT_PCT)
    input_cell(ws, "C13", fcf.get("tax_rate"), fmt=ST.FMT_PCT)
    cell(ws, "A14", "Less: cash taxes")
    cell(ws, "B14", "=-(B11+B12)*B13", fmt=ST.FMT_USD_M)
    cell(ws, "C14", "=-(C11+C12)*C13", fmt=ST.FMT_USD_M)
    cell(ws, "A15", "Less: capital expenditures")
    input_cell(ws, "B15", fcf.get("capex"), fmt=ST.FMT_USD_M)
    input_cell(ws, "C15", fcf.get("capex_y2", fcf.get("capex")), fmt=ST.FMT_USD_M)
    cell(ws, "A16", "Less: working capital & other")
    input_cell(ws, "B16", fcf.get("wc"), fmt=ST.FMT_USD_M)
    input_cell(ws, "C16", fcf.get("wc_y2", fcf.get("wc")), fmt=ST.FMT_USD_M)
    cell(ws, "A17", "Free cash flow")
    cell(ws, "B17", "=B11+B12+B14+B15+B16", fmt=ST.FMT_USD_M, bold=True)
    cell(ws, "C17", "=C11+C12+C14+C15+C16", fmt=ST.FMT_USD_M, bold=True)
    cell(ws, "A18", "FCF per share ($/sh)")
    cell(ws, "B18", f"=B17/Assumptions!$B${A['fd_shares']}", fmt=ST.FMT_USD_2)
    cell(ws, "C18", f"=C17/Assumptions!$B${A['fd_shares']}", fmt=ST.FMT_USD_2)
    cell(ws, "A19", "FCF yield at current price")
    cell(ws, "B19", f"=B18/Assumptions!$B${A['price']}", fmt=ST.FMT_PCT)
    cell(ws, "C19", f"=C18/Assumptions!$B${A['price']}", fmt=ST.FMT_PCT)
    cell(ws, "A20", "FCF conversion (% of adj EBITDA)")
    cell(ws, "B20", "=B17/B11", fmt=ST.FMT_PCT)
    cell(ws, "C20", "=C17/C11", fmt=ST.FMT_PCT)

    section(ws, 22, "LEVERAGE & CAPITAL-RETURN CAPACITY", last_col="C")
    cell(ws, "A23", f"Net debt / {ey} adj EBITDA")
    cell(ws, "B23", f"=Assumptions!B{A['net_spin']}/B11", fmt=ST.FMT_MULT)
    cell(ws, "A24", f"(Net debt + redeemable NCI) / {ey} EBITDA")
    cell(ws, "B24", f"=(Assumptions!B{A['net_spin']}+Assumptions!B{A['red_nci']})/B11", fmt=ST.FMT_MULT)
    cell(ws, "A25", "2-yr cumulative FCF")
    cell(ws, "B25", "=B17+C17", fmt=ST.FMT_USD_M)
    cell(ws, "A26", "  as % of current market cap")
    cell(ws, "B26", f"=B25/Assumptions!B{A['mktcap']}", fmt=ST.FMT_PCT)
    cell(ws, "A28", c["meta"].get("fcf_footnote",
         "Post-spin, the spin entity's EBITDA and interest depart; RemainCo FCF conversion rises."),
         font=ST.FONT_NOTE, wrap=True)
    ws.merge_cells(f"A28:G29")
