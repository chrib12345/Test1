"""Design system extracted from the SPGI master template.

All fills/fonts/number formats below were lifted directly from the uploaded
SPGI_5Segment_SOTP.xlsx so generated workbooks match the original look.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette --------------------------------------------------------------
NAVY = "1F3864"        # section-header band (white bold text on navy)
LIGHT_BLUE = "D9E2F3"  # sub-header / scenario-header band
MED_BLUE = "2E75B6"    # accent band
YELLOW = "FFFF00"      # estimate / refresh-critical input
WHITE = "FFFFFF"

INPUT_BLUE = "0000FF"  # blue font  -> hard input
LINK_GREEN = "008000"  # green font -> cross-sheet link
NOTE_GRAY = "808080"   # 8pt gray   -> source note (col G)
WARN_RED = "C00000"    # red        -> warning / caution

# ---- number formats -------------------------------------------------------
FMT_PCT = r"0.0%;\(0.0%\);\-"
FMT_USD_M = r"\$#,##0;\"($\"#,##0\);\-"          # $ millions
FMT_USD_2 = r"\$#,##0.00;\"($\"#,##0.00\)"        # $ per share (2 dp)
FMT_MULT = r"0.0\x"                                # valuation multiple
FMT_NUM = r"#,##0;\(#,##0\);\-"
FMT_NUM1 = r"#,##0.0"
FMT_PLAIN2 = "0.00"

# ---- reusable style atoms -------------------------------------------------
FONT_BASE = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_INPUT = Font(name="Arial", size=10, color=INPUT_BLUE)
FONT_LINK = Font(name="Arial", size=10, color=LINK_GREEN)
FONT_NOTE = Font(name="Arial", size=8, color=NOTE_GRAY)
FONT_WARN = Font(name="Arial", size=8, color=WARN_RED)
FONT_HDR = Font(name="Arial", size=10, bold=True, color=WHITE)
FONT_TITLE = Font(name="Arial", size=20, bold=True, color=WHITE)

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_MED = PatternFill("solid", fgColor=MED_BLUE)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW)

_thin = Side(style="thin", color="BFBFBF")
BORDER_BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_bold = Side(style="medium", color="000000")
BORDER_BASECASE = Border(left=_bold, right=_bold, top=_bold, bottom=_bold)


def cell(ws, coord, value=None, *, font=None, fill=None, fmt=None,
         align=None, wrap=False, bold=False, border=None):
    """Set a cell's value and style in one call."""
    c = ws[coord]
    if value is not None:
        c.value = value
    if font is not None:
        c.font = font
    elif bold:
        c.font = FONT_BOLD
    else:
        c.font = FONT_BASE
    if fill is not None:
        c.fill = fill
    if fmt is not None:
        c.number_format = fmt
    if border is not None:
        c.border = border
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c


def section(ws, coord_or_row, text, last_col="H"):
    """Navy section-header band spanning columns A..last_col on a row."""
    if isinstance(coord_or_row, int):
        row = coord_or_row
    else:
        row = ws[coord_or_row].row
    a = get_column_letter(1)
    cell(ws, f"{a}{row}", text, font=FONT_HDR, fill=FILL_NAVY)
    for col in range(2, _col_index(last_col) + 1):
        cl = get_column_letter(col)
        ws[f"{cl}{row}"].fill = FILL_NAVY


def _col_index(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def input_cell(ws, coord, value, *, fmt=None, estimate=False):
    """Blue-font hard input; yellow fill when it is an estimate/refresh-critical."""
    return cell(ws, coord, value, font=FONT_INPUT,
                fill=FILL_YELLOW if estimate else None, fmt=fmt)


def note(ws, coord, text):
    """8pt gray source note (column G convention)."""
    return cell(ws, coord, text, font=FONT_NOTE, wrap=True, align="left")


def setup_sheet(ws, tab_color=NAVY, freeze=None, widths=None):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    if freeze:
        ws.freeze_panes = freeze
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
